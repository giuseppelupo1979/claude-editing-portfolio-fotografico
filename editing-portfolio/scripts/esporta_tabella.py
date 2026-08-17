#!/usr/bin/env python3
"""
esporta_tabella.py: calcola i punteggi ponderati e produce la tabella di lavoro.

Legge analisi.json (il contratto descritto in SKILL.md) e scrive:
  tabella.csv   una riga per immagine, con i punteggi per tutte le destinazioni
  tabella.xlsx  lo stesso, in fogli separati (immagini, pesi, sequenza, cluster)

Stampa a schermo le classifiche e i cambi di rango fra destinazioni: sono la parte
piu' informativa, perche' dicono se l'autore fa fotografie singole o progetti.

Uso:
  python3 esporta_tabella.py analisi.json [-o CARTELLA_USCITA]
"""

import argparse
import csv
import json
import os
import sys

ASSI = ["autonomia", "forza", "coerenza", "originalita", "tecnica", "funzione"]

PESI = {
    "libro": {"autonomia": 10, "forza": 20, "coerenza": 25, "originalita": 15, "tecnica": 10, "funzione": 20},
    "mostra": {"autonomia": 25, "forza": 25, "coerenza": 15, "originalita": 15, "tecnica": 20, "funzione": 0},
    "web": {"autonomia": 30, "forza": 25, "coerenza": 15, "originalita": 20, "tecnica": 10, "funzione": 0},
    "concorso": {"autonomia": 25, "forza": 30, "coerenza": 10, "originalita": 25, "tecnica": 10, "funzione": 0},
}


def punteggio(voti, pesi):
    tot = sum(pesi.values())
    acc = 0.0
    for asse, peso in pesi.items():
        if peso == 0:
            continue
        v = voti.get(asse)
        if v is None:
            continue
        acc += float(v) * peso
    return round(acc / tot, 2) if tot else None


def main():
    ap = argparse.ArgumentParser(description="Punteggi ponderati e tabella di lavoro da analisi.json")
    ap.add_argument("analisi")
    ap.add_argument("-o", "--uscita", default=None)
    args = ap.parse_args()

    percorso = os.path.abspath(os.path.expanduser(args.analisi))
    with open(percorso, encoding="utf-8") as fh:
        dati = json.load(fh)
    uscita = os.path.abspath(os.path.expanduser(args.uscita or os.path.dirname(percorso) or "."))
    os.makedirs(uscita, exist_ok=True)

    immagini = dati.get("immagini") or []
    if not immagini:
        sys.exit("analisi.json non contiene immagini.")

    ids = set()
    for im in immagini:
        if not im.get("id"):
            sys.exit("Ogni immagine deve avere un id.")
        if im["id"] in ids:
            sys.exit("Id duplicato: %s" % im["id"])
        ids.add(im["id"])
        mancanti = [a for a in ASSI if im.get("voti", {}).get(a) is None]
        if mancanti:
            print("Attenzione: %s senza voti su %s" % (im["id"], ", ".join(mancanti)))

    for chiave, elenco in (
        ("sequenza.spread", [x for coppia in (dati.get("sequenza", {}).get("spread") or []) for x in coppia]),
        ("cluster", [i for c in (dati.get("cluster") or []) for i in c.get("immagini", [])]),
        ("gallerie", [i for g in (dati.get("gallerie") or []) for i in g.get("ordine", [])]),
        ("scarti", [s.get("id") for s in (dati.get("scarti") or [])]),
    ):
        fantasmi = sorted({x for x in elenco if x and x not in ids})
        if fantasmi:
            sys.exit("In %s ci sono id inesistenti: %s" % (chiave, ", ".join(fantasmi)))

    for im in immagini:
        voti = im.get("voti", {})
        im["_p"] = {d: punteggio(voti, p) for d, p in PESI.items()}

    ranghi = {}
    for d in PESI:
        ordinati = sorted(immagini, key=lambda i: (i["_p"][d] is None, -(i["_p"][d] or 0)))
        ranghi[d] = {im["id"]: k + 1 for k, im in enumerate(ordinati)}

    dest = (dati.get("progetto") or {}).get("destinazione") or "libro"
    if dest not in PESI:
        dest = "libro"

    righe = []
    for im in immagini:
        voti = im.get("voti", {})
        r = {
            "id": im["id"],
            "file": im.get("file", ""),
            "titolo": im.get("titolo", ""),
            "genere": im.get("genere", ""),
            "verdetto": im.get("verdetto", ""),
            "ruolo": im.get("ruolo", ""),
            "cluster": " | ".join(im.get("cluster", []) or []),
            "cieco": im.get("cieco", ""),
            "registro_tonale": im.get("registro_tonale", ""),
            "densita": im.get("densita", ""),
        }
        for a in ASSI:
            r["v_" + a] = voti.get(a, "")
        for d in PESI:
            r["p_" + d] = im["_p"][d]
            r["r_" + d] = ranghi[d][im["id"]]
        r["scarto_rango_libro_concorso"] = ranghi["concorso"][im["id"]] - ranghi["libro"][im["id"]]
        r["forza_principale"] = im.get("forza_principale", "")
        r["limite_principale"] = im.get("limite_principale", "")
        r["didascalia"] = im.get("didascalia", "")
        righe.append(r)

    righe.sort(key=lambda r: r["r_" + dest])

    csv_out = os.path.join(uscita, "tabella.csv")
    with open(csv_out, "w", newline="", encoding="utf-8") as fh:
        wr = csv.DictWriter(fh, fieldnames=list(righe[0].keys()))
        wr.writeheader()
        wr.writerows(righe)

    xlsx_out = None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "immagini"
        intestazioni = list(righe[0].keys())
        ws.append(intestazioni)
        for r in righe:
            ws.append([r[c] for c in intestazioni])
        grassetto = Font(bold=True, color="FFFFFF")
        sfondo = PatternFill("solid", fgColor="2F3B52")
        for c in range(1, len(intestazioni) + 1):
            cella = ws.cell(row=1, column=c)
            cella.font = grassetto
            cella.fill = sfondo
            cella.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "C2"
        larghezze = {"id": 8, "file": 26, "titolo": 22, "cluster": 20, "forza_principale": 48,
                     "limite_principale": 48, "didascalia": 42, "verdetto": 15, "ruolo": 12}
        for idx, nome in enumerate(intestazioni, 1):
            ws.column_dimensions[get_column_letter(idx)].width = larghezze.get(nome, 11)

        wp = wb.create_sheet("pesi")
        wp.append(["destinazione"] + ASSI)
        for d, p in PESI.items():
            wp.append([d] + [p[a] for a in ASSI])
        for c in range(1, len(ASSI) + 2):
            wp.cell(row=1, column=c).font = Font(bold=True)

        seq = dati.get("sequenza", {}).get("spread") or []
        if seq:
            wsq = wb.create_sheet("sequenza")
            wsq.append(["spread", "pagina sinistra", "pagina destra", "ruolo sx", "ruolo dx"])
            per_id = {im["id"]: im for im in immagini}
            for n, coppia in enumerate(seq, 1):
                sx, dx = (list(coppia) + [None, None])[:2]
                wsq.append([
                    n, sx or "(bianca)", dx or "(bianca)",
                    (per_id.get(sx or "", {}) or {}).get("ruolo", ""),
                    (per_id.get(dx or "", {}) or {}).get("ruolo", ""),
                ])
            for c in range(1, 6):
                wsq.cell(row=1, column=c).font = Font(bold=True)

        cl = dati.get("cluster") or []
        if cl:
            wc = wb.create_sheet("cluster")
            wc.append(["id", "nome", "registro_nome", "forza", "n_immagini", "immagini", "tesi"])
            for c in cl:
                wc.append([
                    c.get("id", ""), c.get("nome", ""), c.get("registro_nome", ""),
                    c.get("forza", ""), len(c.get("immagini", [])),
                    " ".join(c.get("immagini", [])), c.get("tesi", ""),
                ])
            for i in range(1, 8):
                wc.cell(row=1, column=i).font = Font(bold=True)
            wc.column_dimensions["G"].width = 60

        xlsx_out = os.path.join(uscita, "tabella.xlsx")
        wb.save(xlsx_out)
    except ImportError:
        print("openpyxl assente: prodotto solo il CSV. pip install openpyxl --break-system-packages")

    print("Destinazione dichiarata: %s" % dest)
    print("Immagini: %d" % len(immagini))
    print("\nPrime 10 per la destinazione dichiarata (%s):" % dest)
    for r in righe[:10]:
        print("  %2d. %s  %.2f  %-14s %s" % (
            r["r_" + dest], r["id"], r["p_" + dest] or 0, r["verdetto"] or "-", r["file"][:34]))

    movimenti = sorted(righe, key=lambda r: -abs(r["scarto_rango_libro_concorso"]))[:6]
    print("\nCambi di rango piu' forti fra libro e concorso (il dato che dice che tipo di autore e'):")
    for r in movimenti:
        s = r["scarto_rango_libro_concorso"]
        verso = "meglio nel libro" if s > 0 else "meglio al concorso"
        print("  %s  libro #%d, concorso #%d  (%+d, %s)" % (
            r["id"], r["r_libro"], r["r_concorso"], s, verso))

    conteggio = {}
    for r in righe:
        conteggio[r["verdetto"] or "senza verdetto"] = conteggio.get(r["verdetto"] or "senza verdetto", 0) + 1
    print("\nVerdetti: " + ", ".join("%s %d" % (k, v) for k, v in sorted(conteggio.items(), key=lambda kv: -kv[1])))

    scarti = dati.get("scarti") or []
    print("Scarti dichiarati: %d" % len(scarti))
    print("\nScritto: %s" % csv_out)
    if xlsx_out:
        print("Scritto: %s" % xlsx_out)


if __name__ == "__main__":
    main()
